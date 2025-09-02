# Import error logging utilities for error tracking
from utils.error_logger import capture_exception, capture_message, set_tag, set_context

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from datetime import datetime
import uuid
import logging
from typing import Optional, List, Dict

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def save_test_script(content: str, base_name: str) -> Optional[str]:
    """Save test script to file.

    Args:
        content (str): Content to write to file
        base_name (str): Base name for the file

    Returns:
        Optional[str]: Filename if successful, None otherwise
    """
    if not base_name or not content:
        logger.error("Filename and content cannot be empty")
        print("❌ Filename and content cannot be empty")
        return None

    output_dir = os.path.join("tests", "generated")
    # Ensure the filename is valid
    filename = f"{base_name}.txt"
    file_path = os.path.join(output_dir, filename)

    try:
        os.makedirs(output_dir, exist_ok=True)
        with open(file_path, "w", encoding="utf-8") as file:
            file.write(content)
        logger.info(f"Test script saved successfully: {filename}")
        return filename
    except Exception as e:
        logger.error(f"Error saving file {filename}: {e}")
        print(f"❌ Error saving file {filename}: {e}")
        return None

def save_excel_report(test_cases: str, base_name: str) -> Optional[str]:
    """Save test cases to Excel file.

    Args:
        test_cases (str): Test cases content to write to Excel
        base_name (str): Base name for the file

    Returns:
        Optional[str]: Filename if successful, None otherwise
    """
    if not base_name or not test_cases:
        logger.error("Filename and test cases cannot be empty")
        print("❌ Filename and test cases cannot be empty")
        return None

    filename = f"{base_name}.xlsx"
    output_dir = os.path.join("tests", "generated")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    try:
        # First, check for TEST TYPE section markers
        sections = extract_test_type_sections(test_cases)
        
        # If no explicit TEST TYPE sections found, use traditional parsing
        if not sections:
            logger.info("No TEST TYPE sections found, using traditional parsing")
            test_data = parse_traditional_format(test_cases)
        else:
            # Parse each section separately
            logger.info(f"Found {len(sections)} TEST TYPE sections")
            test_data = []
            for section_name, section_content in sections.items():
                section_data = parse_traditional_format(section_content, default_section=section_name)
                test_data.extend(section_data)

        # Convert to DataFrame
        df = pd.DataFrame(test_data) if test_data else pd.DataFrame()
        
        if df.empty:
            logger.warning("No test cases could be parsed")
            print("⚠️ No test cases could be parsed")
            # Create a minimal DataFrame to avoid errors
            df = pd.DataFrame({
                'Section': ['No test cases found'],
                'Title': ['No test cases could be parsed'],
                'Scenario': [''],
                'Steps': [''],
                'Expected Result': [''],
                'Actual Result': ['']
            })
        
        # Fill empty values with appropriate defaults
        if 'Section' not in df.columns:
            df['Section'] = 'General'
        else:
            df['Section'] = df['Section'].fillna('General')
            
        # Ensure all required columns exist
        required_columns = ['Title', 'Scenario', 'Steps', 'Expected Result', 'Actual Result', 'Status', 'Priority']
        for col in required_columns:
            if col not in df.columns:
                df[col] = ''
                
        # Convert list of steps to string if needed        
        if 'Steps' in df.columns:            
            df['Steps'] = df['Steps'].apply(lambda x: '\n'.join([f"{i+1}. {step}" for i, step in enumerate(x)]) if isinstance(x, list) else (x or ''))
        
        # Fill all remaining NaN values
        df = df.fillna('')

        # Reorder columns, starting with the most important ones
        column_order = ['Section', 'Title', 'Scenario', 'Steps', 'Expected Result', 'Status', 'Actual Result', 'Priority']
        # Add any additional columns that might be present
        for col in df.columns:
            if col not in column_order:
                column_order.append(col)
                
        # Keep only columns that exist in the DataFrame
        column_order = [col for col in column_order if col in df.columns]
        df = df.reindex(columns=column_order)

        # Save to Excel with improved formatting
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Test Cases')
            worksheet = writer.sheets['Test Cases']
            
            # Adjust column widths
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(col)
                )
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, 50)

        logger.info(f"Excel report saved successfully: {filename}")
        return filename

    except Exception as e:
        logger.error(f"Error saving Excel report: {e}")
        print(f"❌ Error saving Excel report: {e}")
        return None

def create_excel_report(test_cases: List[Dict], status_values: Dict, source_type: str, item_ids: List[str]) -> bytes:
    """Create an Excel report from test cases data.
    
    Args:
        test_cases (List[Dict]): List of test case dictionaries
        status_values (Dict): Status values for test cases
        source_type (str): Source type (e.g., 'Jira', 'Azure')
        item_ids (List[str]): List of item IDs
        
    Returns:
        bytes: Excel file as bytes
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Create a new workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        summary_ws = wb.create_sheet("Summary")
        
        # Add summary information
        summary_ws['A1'] = f"Test Cases Report - {source_type}"
        summary_ws['A1'].font = Font(bold=True, size=16)
        
        summary_ws['A3'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        summary_ws['A4'] = f"Source Type: {source_type}"
        summary_ws['A5'] = f"Item IDs: {', '.join(item_ids)}"
        summary_ws['A6'] = f"Total Test Cases: {len(test_cases)}"
        
        # Add status summary
        status_counts = {}
        for tc in test_cases:
            tc_id = tc.get('Title', 'Unknown')
            status = status_values.get(tc_id, 'Not Tested')
            status_counts[status] = status_counts.get(status, 0) + 1
        
        summary_ws['A8'] = "Status Summary:"
        summary_ws['A8'].font = Font(bold=True)
        
        row = 9
        for status, count in status_counts.items():
            summary_ws[f'A{row}'] = f"{status}: {count}"
            row += 1
        
        # Create test cases sheet
        tc_ws = wb.create_sheet("Test Cases")
        
        # Define headers
        headers = ['Title', 'Scenario', 'Steps', 'Expected Result', 'Status', 'Item ID']
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = tc_ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add test cases data
        for row, tc in enumerate(test_cases, 2):
            # Extract item ID from title (e.g., "Title (KAN-6)" -> "KAN-6")
            title = tc.get('Title', '')
            item_id = ''
            if '(' in title and ')' in title:
                item_id = title.split('(')[-1].split(')')[0]
            
            tc_ws.cell(row=row, column=1, value=tc.get('Title', ''))
            tc_ws.cell(row=row, column=2, value=tc.get('Scenario', ''))
            
            # Handle Steps field - convert list to string if needed
            steps = tc.get('Steps', '')
            if isinstance(steps, list):
                steps = '\n'.join([f"{i+1}. {step}" for i, step in enumerate(steps)])
            tc_ws.cell(row=row, column=3, value=steps)
            
            # Handle Expected Result field - convert list to string if needed
            expected_result = tc.get('Expected Result', '')
            if isinstance(expected_result, list):
                expected_result = '\n'.join(expected_result)
            tc_ws.cell(row=row, column=4, value=expected_result)
            
            # Get status from status_values
            tc_id = tc.get('Title', '')
            status = status_values.get(tc_id, 'Not Tested')
            tc_ws.cell(row=row, column=5, value=status)
            
            tc_ws.cell(row=row, column=6, value=item_id)
        
        # Auto-adjust column widths
        for column in tc_ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            tc_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
        
    except Exception as e:
        logger.error(f"Error creating Excel report: {e}")
        raise e

def extract_test_type_sections(test_cases: str) -> Dict[str, str]:
    """Extract sections from test cases content based on TEST TYPE markers.
    
    Args:
        test_cases (str): The full test cases content
        
    Returns:
        Dict[str, str]: Dictionary of section name to section content
    """
    sections = {}
    
    # Look for TEST TYPE: section markers
    section_pattern = r"TEST TYPE:\s*([^\n]+)"
    section_matches = list(re.finditer(section_pattern, test_cases))
    
    # If no TEST TYPE markers found, return empty dict
    if not section_matches:
        return {}
    
    # Process each section
    for i, match in enumerate(section_matches):
        section_name = match.group(1).strip()
        start_pos = match.end()
        # End position is either the start of next section or end of content
        end_pos = section_matches[i+1].start() if i+1 < len(section_matches) else len(test_cases)
        # Extract section content
        section_content = test_cases[start_pos:end_pos].strip()
        sections[section_name] = section_content
        
    return sections

def parse_traditional_format(test_cases: str, default_section: str = "General") -> List[Dict]:
    """Parse test cases using the traditional format.
    
    Args:
        test_cases (str): Test cases content to parse
        default_section (str): Default section name if none is specified
        
    Returns:
        List[Dict]: List of test case dictionaries
    """
    test_data = []
    current_test = {}
    current_section = default_section
    collecting_steps = False
    current_steps = []

    logger.info(f"Parsing test cases with default section: {default_section}")
    
    # Special handling for test cases without clear delimiters - try to extract full test case blocks
    if ("Title:" in test_cases and "Steps to reproduce:" in test_cases) or ("Steps to reproduce:" in test_cases and re.search(r'TC_[A-Z]+_\d+', test_cases)):
        logger.info("Detected standard test case format with Title and Steps to reproduce")
        # Split by blank lines to find test case boundaries
        # This handles cases where test cases are separated by blank lines
        test_case_blocks = re.split(r'\n\s*\n', test_cases)
        for block in test_case_blocks:
            block = block.strip()
            if not block or len(block) < 10:  # Skip empty or very short blocks
                continue
                
            # Check if this looks like a test case - either has Title: or starts with TC_ pattern
            if "Title:" in block or re.match(r'TC_[A-Z]+_\d+', block.strip()):
                test_case = {}
                test_case['Section'] = default_section
                
                # Extract title - handle both "Title:" format and direct TC_ pattern
                title_match = re.search(r'Title:\s*(.*?)(?:\n|$)', block)
                if title_match:
                    test_case['Title'] = title_match.group(1).strip().replace('**', '').strip()
                else:
                    # Try to extract title from the first line if it starts with TC_ pattern
                    first_line = block.split('\n')[0].strip()
                    if re.match(r'TC_[A-Z]+_\d+', first_line):
                        test_case['Title'] = first_line.replace('**', '').strip()
                
                # Extract scenario
                scenario_match = re.search(r'Scenario:\s*(.*?)(?:\n|$)', block)
                if scenario_match:
                    test_case['Scenario'] = scenario_match.group(1).strip().replace('**', '').strip()
                
                # Extract steps - improved regex to handle multi-line steps
                steps_match = re.search(r'(?:\*\*)?Steps to reproduce:(?:\*\*)?\s*\n([\s\S]*?)(?=\n\s*(?:\*\*)?Expected Result:|$)', block, re.DOTALL)
                if steps_match:
                    steps_text = steps_match.group(1).strip()
                    logger.info(f"Found steps text: {steps_text[:200]}...")
                    # Try to split steps by numbered lines or bullet points
                    step_lines = re.findall(r'(?:^|\n)\s*(?:\d+\.\s*|\*\s*|\-\s*)(.*?)(?=\n\s*(?:\d+\.\s*|\*\s*|\-\s*)|$)', steps_text, re.MULTILINE)
                    if step_lines:
                        test_case['Steps'] = [step.strip().replace('**', '').strip() for step in step_lines if step.strip()]
                        logger.info(f"Extracted {len(test_case['Steps'])} steps using numbered format")
                    else:
                        # If no clear step delimiters, split by newlines and filter out empty lines
                        steps = [step.strip().replace('**', '').strip() for step in steps_text.split('\n') if step.strip()]
                        test_case['Steps'] = steps
                        logger.info(f"Extracted {len(test_case['Steps'])} steps using line-by-line format")
                else:
                    logger.warning(f"No steps found in block: {block[:200]}...")
                
                # Extract expected result (stop before Actual Result, Status, or Priority)
                expected_match = re.search(r'Expected Result:\s*(.*?)(?:\n\s*Actual Result:|\n\s*Status:|\n\s*Priority:|$)', block, re.DOTALL)
                if expected_match:
                    test_case['Expected Result'] = expected_match.group(1).strip().replace('**', '').strip()
                
                # Extract status if present as its own line
                status_match = re.search(r'\n\s*Status:\s*(.*?)(?:\n|$)', block)
                if status_match:
                    test_case['Status'] = status_match.group(1).strip()
                
                # Extract actual result if present
                actual_match = re.search(r'Actual Result:\s*(.*?)(?:\n\s*Priority:|$)', block, re.DOTALL)
                if actual_match:
                    actual_result = actual_match.group(1).strip()
                    # Only set if it's not empty and doesn't contain "Priority:"
                    if actual_result and not actual_result.startswith('Priority:'):
                        test_case['Actual Result'] = actual_result.replace('**', '').strip()
                
                # Extract priority if present
                priority_match = re.search(r'Priority:\s*(.*?)(?:\n|$)', block)
                if priority_match:
                    test_case['Priority'] = priority_match.group(1).strip().replace('**', '').strip()
                
                # Only add test case if it has a title
                if test_case.get('Title'):
                    test_data.append(test_case)
                    logger.debug(f"Extracted test case: {test_case['Title']}")
                
        if test_data:
            logger.info(f"Extracted {len(test_data)} test cases using block parsing")
            return test_data
    
    # If block parsing didn't work, fall back to line-by-line parsing
    logger.info("Using line-by-line parsing for test cases")
    
    # Split by lines for easier processing
    lines = test_cases.split('\n')
    i = 0
    
    while i < len(lines):
        line = lines[i].strip()
        i += 1
        
        if not line:
            continue
            
        # Check for section headers (both markdown and plain text)
        if line.startswith('###'):
            current_section = line.replace('#', '').strip() or default_section
            continue
            
        # Handle various title formats
        title_match = re.match(r'^(?:\d+\.\s*)?(?:\*\*)?Title:(?:\*\*)?\s*(.*?)$', line)
        if title_match:
            # Save previous test case if exists
            if current_test:
                if collecting_steps:
                    current_test['Steps'] = current_steps
                test_data.append(current_test)
                
            # Start a new test case
            current_test = {
                'Section': current_section,
                'Title': title_match.group(1).strip().replace('**', '').strip()
            }
            collecting_steps = False
            current_steps = []
            continue
            
        # Handle scenario
        scenario_match = re.match(r'^(?:\*\*)?Scenario:(?:\*\*)?\s*(.*?)$', line)
        if scenario_match and current_test:
            current_test['Scenario'] = scenario_match.group(1).strip().replace('**', '').strip()
            continue
            
        # Handle steps header - support multiple variations
        steps_match = re.match(r'^(?:\*\*)?Steps(?: to reproduce)?:(?:\*\*)?', line)
        if steps_match and current_test:
            collecting_steps = True
            current_steps = []
            logger.info(f"Started collecting steps for test case: {current_test.get('Title', 'Unknown')}")
            continue
            
        # Collect steps
        if collecting_steps:
            # Check if we're now on the Expected Result section
            if re.match(r'^(?:\*\*)?Expected Result:(?:\*\*)?', line):
                collecting_steps = False
                current_test['Steps'] = current_steps
                logger.info(f"Finished collecting {len(current_steps)} steps for test case: {current_test.get('Title', 'Unknown')}")
                
                # Extract expected result value
                er_match = re.match(r'^(?:\*\*)?Expected Result:(?:\*\*)?\s*(.*?)$', line)
                if er_match:
                    current_test['Expected Result'] = er_match.group(1).strip().replace('**', '').strip()
                continue
                
            # Support various step formats (1. Step, - Step, * Step, etc.)
            step_match = re.match(r'^(?:(\d+)\.|\-|\*)\s*(.*?)$', line)
            if step_match:
                step_text = step_match.group(2) if step_match.groups()[-1] else line
                # Clean up markdown formatting from steps
                cleaned_step = step_text.strip().replace('**', '').strip()
                if cleaned_step:  # Only add non-empty steps
                    current_steps.append(cleaned_step)
                    logger.debug(f"Added step: {cleaned_step}")
                continue
            elif line.strip():  # If not empty and doesn't match step pattern, add as-is
                cleaned_line = line.strip().replace('**', '').strip()
                if cleaned_line:  # Only add non-empty lines
                    current_steps.append(cleaned_line)
                    logger.debug(f"Added step (no pattern): {cleaned_line}")
                continue
                
        # Handle expected result
        expected_match = re.match(r'^(?:\*\*)?Expected Result:(?:\*\*)?\s*(.*?)$', line)
        if expected_match and current_test:
            current_test['Expected Result'] = expected_match.group(1).strip().replace('**', '').strip()
            continue

        # Handle status as standalone line
        status_alone = re.match(r'^Status:\s*(.*?)$', line)
        if status_alone and current_test:
            current_test['Status'] = status_alone.group(1).strip()
            continue
            
        # Handle actual result
        actual_match = re.match(r'^(?:\*\*)?Actual Result:(?:\*\*)?\s*(.*?)$', line)
        if actual_match and current_test:
            current_test['Actual Result'] = actual_match.group(1).strip().replace('**', '').strip()
            continue
            
        # Handle priority
        priority_match = re.match(r'^(?:\*\*)?Priority:(?:\*\*)?\s*(.*?)$', line)
        if priority_match and current_test:
            current_test['Priority'] = priority_match.group(1).strip().replace('**', '').strip()
            continue
            
        # If we're not collecting steps and this is a continuation line, append to the last field
        if current_test and not collecting_steps and line:
            # Try to determine which field this continuation belongs to
            if 'Priority' in current_test:
                current_test['Priority'] += ' ' + line
            elif 'Actual Result' in current_test:
                current_test['Actual Result'] += ' ' + line
            elif 'Expected Result' in current_test:
                current_test['Expected Result'] += ' ' + line
            elif 'Scenario' in current_test:
                current_test['Scenario'] += ' ' + line

    # Add the last test case if it exists
    if current_test:
        if collecting_steps:
            current_test['Steps'] = current_steps
        test_data.append(current_test)
        
    logger.info(f"Extracted {len(test_data)} test cases using line-by-line parsing")
    return test_data